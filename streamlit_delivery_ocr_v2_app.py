import os
import re
import tempfile
from io import BytesIO
from difflib import get_close_matches

import pandas as pd
import pytesseract
import streamlit as st
from PIL import Image, ImageOps, ImageFilter
from pdf2image import convert_from_path
from openpyxl import load_workbook


st.set_page_config(page_title="PDF OCR Tracker Updater", layout="wide")
st.title("PDF OCR Tracker Updater")


# -----------------------------
# Session State
# -----------------------------
if "processed" not in st.session_state:
    st.session_state.processed = False

if "results" not in st.session_state:
    st.session_state.results = {}

if "run_id" not in st.session_state:
    st.session_state.run_id = 0


# -----------------------------
# Normalization Helpers
# -----------------------------
def normalize_text(value):
    if value is None:
        return ""
    value = str(value).strip()
    if value.startswith("'"):
        value = value[1:]
    value = value.replace("\xa0", " ")
    value = " ".join(value.split())
    return value.upper()


def normalize_header(value):
    if value is None:
        return ""
    return " ".join(str(value).strip().lower().split())


def normalize_part_number(value):
    if value is None:
        return ""
    value = str(value).strip()
    if value.startswith("'"):
        value = value[1:]
    value = value.replace("\xa0", " ")
    value = value.replace("—", "-").replace("–", "-")
    value = " ".join(value.split())
    return value.upper()


def normalize_ocr_part_text(value):
    if value is None:
        return ""

    value = str(value).strip()
    value = value.replace("—", "-").replace("–", "-")
    value = value.upper()

    # Common OCR substitutions
    value = value.replace("O", "0")
    value = value.replace("I", "1")
    value = value.replace("L", "1")

    value = re.sub(r"\s*-\s*", "-", value)
    value = re.sub(r"\s+", "", value)

    if re.fullmatch(r"\d{8}", value):
        value = value[:4] + "-" + value[4:]

    return value


def combine_unique_values(series):
    values = []
    for v in series:
        if pd.isna(v):
            continue
        v = str(v).strip()
        if v != "":
            values.append(v)

    unique_values = sorted(set(values), key=lambda x: (len(x), x))
    return ", ".join(unique_values)


def merge_comma_separated(existing_value, new_value):
    existing_items = []
    new_items = []

    if existing_value not in [None, ""]:
        existing_items = [x.strip() for x in str(existing_value).split(",") if x.strip()]

    if new_value not in [None, ""]:
        new_items = [x.strip() for x in str(new_value).split(",") if x.strip()]

    merged = sorted(set(existing_items + new_items), key=lambda x: (len(x), x))
    return ", ".join(merged)


def normalize_quantity_text(qty_raw):
    qty_raw = str(qty_raw).strip().replace(" ", "")

    if "," in qty_raw and "." in qty_raw:
        if qty_raw.rfind(",") > qty_raw.rfind("."):
            qty_clean = qty_raw.replace(".", "").replace(",", ".")
        else:
            qty_clean = qty_raw.replace(",", "")
    elif "," in qty_raw:
        parts = qty_raw.split(",")
        if len(parts[-1]) == 3 and len(parts) > 1:
            qty_clean = qty_raw.replace(",", "")
        else:
            qty_clean = qty_raw.replace(",", ".")
    elif "." in qty_raw:
        parts = qty_raw.split(".")
        if len(parts[-1]) == 3 and len(parts) > 1:
            qty_clean = qty_raw.replace(".", "")
        else:
            qty_clean = qty_raw
    else:
        qty_clean = qty_raw

    return int(float(qty_clean))


# -----------------------------
# OCR / Parsing Helpers
# -----------------------------
def extract_document_number(text):
    patterns = [
        r"Truck\s*No\.?\s*[:#]?\s*([A-Z0-9\-\/]+)",
        r"Truck\s*[:#]?\s*([A-Z0-9\-\/]+)",
    ]

    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return match.group(1).strip()

    return ""


def is_table_header(line):
    line_lower = line.lower()

    item_words = ["item", "artikelnr", "artikelnr."]
    desc_words = ["descr", "beschr", "beschr."]

    has_item = any(word in line_lower for word in item_words)
    has_desc = any(word in line_lower for word in desc_words)

    return has_item and has_desc


def is_end_of_table(line):
    line_lower = line.lower().strip()
    end_words = ["total colli", "anzahl colli", "anzahl colli:"]
    return any(word in line_lower for word in end_words)


def extract_colli_number(line):
    patterns = [
        r"colli\s*#?\s*([0-9]+)",
        r"colli\s*nr\.?\s*[:#]?\s*([0-9]+)",
    ]

    for pattern in patterns:
        match = re.search(pattern, line, re.IGNORECASE)
        if match:
            return match.group(1).strip()

    return ""


def preprocess_for_ocr(img):
    img = ImageOps.exif_transpose(img)
    img = img.convert("L")
    img = ImageOps.autocontrast(img)

    w, h = img.size
    img = img.resize((w * 2, h * 2))

    img = img.filter(ImageFilter.MedianFilter(size=3))
    img = img.filter(ImageFilter.SHARPEN)

    return img


def run_ocr(img):
    processed_img = preprocess_for_ocr(img)
    text = pytesseract.image_to_string(
        processed_img,
        config="--oem 3 --psm 6 -c preserve_interword_spaces=1"
    )
    return processed_img, text


def line_looks_like_item_row(line):
    line = line.replace("—", "-").replace("–", "-")
    patterns = [
        r"\b\d{4}\s*-\s*\d{4}\b",
        r"\b\d{4}\s+\d{4}\b",
        r"\b\d{8}\b",
    ]
    return any(re.search(pattern, line, re.IGNORECASE) for pattern in patterns)


def extract_item_and_remainder(line):
    cleaned = line.replace("—", "-").replace("–", "-")

    patterns = [
        r"(\d{4}\s*-\s*\d{4})",
        r"(\d{4}\s+\d{4})",
        r"(\d{8})",
    ]

    for pattern in patterns:
        m = re.search(pattern, cleaned, re.IGNORECASE)
        if m:
            raw_item = m.group(1)
            item_no = normalize_ocr_part_text(raw_item)
            remainder = cleaned[m.end():].strip()
            return item_no, remainder

    return None, None


def parse_quantity_and_description(remainder):
    remainder = " ".join(str(remainder).split())

    qty_match = re.search(
        r"(.+?)\s+([\d.,]+)\s+(piece|stück|pcs?|bundle|kg|pc|roll|rolle|set|sets|meter|metre|each|einh\.?|stk|stck|m)\b",
        remainder,
        re.IGNORECASE,
    )
    if qty_match:
        description = qty_match.group(1).strip(" -.;:")
        qty_raw = qty_match.group(2).strip()
        unit = qty_match.group(3).strip().lower()
        try:
            quantity = normalize_quantity_text(qty_raw)
            return description, quantity, unit
        except Exception:
            pass

    fallback_match = re.search(
        r"(.+?)\s+([\d.,]+)(?:\s+\S+)?$",
        remainder,
        re.IGNORECASE,
    )
    if fallback_match:
        description = fallback_match.group(1).strip(" -.;:")
        qty_raw = fallback_match.group(2).strip()
        try:
            quantity = normalize_quantity_text(qty_raw)
            return description, quantity, "auto"
        except Exception:
            pass

    candidates = re.findall(r"[\d.,]+", remainder)
    if candidates:
        for qty_raw in reversed(candidates):
            try:
                quantity = normalize_quantity_text(qty_raw)
                qty_pos = remainder.rfind(qty_raw)
                description = remainder[:qty_pos].strip(" -.;:")
                if description:
                    return description, quantity, "fallback"
            except Exception:
                continue

    return None, None, None


# -----------------------------
# File Loading
# -----------------------------
def load_pages_from_upload(uploaded_file):
    ext = os.path.splitext(uploaded_file.name)[1].lower()

    if ext == ".pdf":
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
            tmp.write(uploaded_file.getvalue())
            tmp_path = tmp.name

        try:
            pages = convert_from_path(tmp_path, dpi=250)
            return [page.convert("RGB") for page in pages]
        finally:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
    else:
        img = Image.open(uploaded_file).convert("RGB")
        return [img]


# -----------------------------
# Tracker Lookup
# -----------------------------
def build_tracker_lookup(wb):
    tracker_rows = {}
    known_tracker_parts = set()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = {}

        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=1, column=col).value
            if value is not None:
                headers[normalize_header(value)] = col

        item_col = headers.get("item #") or headers.get("part #")
        qty_col = headers.get("qty received")
        pallet_col = headers.get("pallet #")
        container_col = headers.get("container")

        if not item_col or not qty_col:
            continue

        for row in range(2, ws.max_row + 1):
            item_value = ws.cell(row=row, column=item_col).value
            if item_value is None:
                continue

            normalized_item = normalize_text(item_value)

            if normalized_item:
                known_tracker_parts.add(normalized_item)

                if normalized_item not in tracker_rows:
                    tracker_rows[normalized_item] = []

                tracker_rows[normalized_item].append(
                    {
                        "sheet": ws,
                        "row": row,
                        "qty_col": qty_col,
                        "pallet_col": pallet_col,
                        "container_col": container_col,
                    }
                )

    return tracker_rows, sorted(known_tracker_parts)


def recover_part_from_tracker(item_no, known_tracker_parts):
    if not item_no:
        return item_no

    item_no = normalize_ocr_part_text(item_no)

    if item_no in known_tracker_parts:
        return item_no

    compact = item_no.replace("-", "")
    for known in known_tracker_parts:
        if known.replace("-", "") == compact:
            return known

    matches = get_close_matches(item_no, known_tracker_parts, n=1, cutoff=0.90)
    if matches:
        return matches[0]

    return item_no


# -----------------------------
# OCR Processing
# -----------------------------
def process_delivery_files(delivery_files, known_tracker_parts):
    all_items = []
    preview_images = []
    ocr_text_records = []
    skipped_line_records = []

    progress = st.progress(0, text="Starting OCR processing...")
    total_files = len(delivery_files)

    for file_index, uploaded_file in enumerate(delivery_files, start=1):
        progress.progress(
            int(((file_index - 1) / total_files) * 100),
            text=f"Processing file {file_index} of {total_files}: {uploaded_file.name}",
        )

        pages = load_pages_from_upload(uploaded_file)
        table_started_in_previous_page = False

        for page_index, page_img in enumerate(pages, start=1):
            processed_img, text = run_ocr(page_img)
            preview_images.append((f"{uploaded_file.name} - Page {page_index}", processed_img))

            ocr_text_records.append(
                {
                    "SourceFile": uploaded_file.name,
                    "PageNumber": page_index,
                    "OCR_Text": text,
                }
            )

            document_number = extract_document_number(text)
            lines = text.split("\n")
            capture = table_started_in_previous_page
            current_colli = ""
            page_item_count = 0

            for line_no, line in enumerate(lines, start=1):
                original_line = line
                line = " ".join(str(line).split())

                if not line:
                    continue

                detected_colli = extract_colli_number(line)
                if detected_colli:
                    current_colli = detected_colli

                if is_table_header(line):
                    capture = True
                    continue

                if not capture and line_looks_like_item_row(line):
                    capture = True

                if not capture:
                    continue

                if is_end_of_table(line):
                    capture = False
                    continue

                line = re.split(
                    r"(total\s+colli|anzahl\s+colli\s*:?)",
                    line,
                    flags=re.IGNORECASE
                )[0].strip()

                if not line:
                    continue

                item_no, remainder = extract_item_and_remainder(line)
                if not item_no:
                    if re.search(r"\d{4}", line):
                        skipped_line_records.append(
                            {
                                "SourceFile": uploaded_file.name,
                                "PageNumber": page_index,
                                "LineNumber": line_no,
                                "Reason": "No item match",
                                "Line": original_line,
                                "CleanedLine": line,
                            }
                        )
                    continue

                item_no = recover_part_from_tracker(item_no, known_tracker_parts)

                description, quantity, unit = parse_quantity_and_description(remainder)
                if quantity is None or description is None:
                    skipped_line_records.append(
                        {
                            "SourceFile": uploaded_file.name,
                            "PageNumber": page_index,
                            "LineNumber": line_no,
                            "Reason": "No quantity/description parse",
                            "Line": original_line,
                            "CleanedLine": line,
                            "ItemNoGuess": item_no,
                            "Remainder": remainder,
                        }
                    )
                    continue

                all_items.append(
                    {
                        "ItemNo": item_no,
                        "Description": description,
                        "Quantity": quantity,
                        "Unit": unit,
                        "ColliNo": current_colli,
                        "DocumentNumber": document_number,
                        "SourceFile": uploaded_file.name,
                        "PageNumber": page_index,
                    }
                )

                page_item_count += 1

            table_started_in_previous_page = page_item_count > 0

    progress.progress(100, text="OCR processing complete.")

    raw_df = pd.DataFrame(all_items)
    ocr_text_df = pd.DataFrame(ocr_text_records)
    skipped_lines_df = pd.DataFrame(skipped_line_records)

    if raw_df.empty:
        summary_df = pd.DataFrame(columns=["ItemNo", "Description", "Quantity", "PalletList", "DocumentList"])
    else:
        summary_df = (
            raw_df.groupby("ItemNo")
            .agg(
                Quantity=("Quantity", "sum"),
                Description=("Description", lambda x: x.mode().iat[0] if not x.mode().empty else x.iloc[0]),
                PalletList=("ColliNo", combine_unique_values),
                DocumentList=("DocumentNumber", combine_unique_values),
            )
            .reset_index()
        )

        summary_df = summary_df[["ItemNo", "Description", "Quantity", "PalletList", "DocumentList"]]

    return raw_df, summary_df, preview_images, ocr_text_df, skipped_lines_df


# -----------------------------
# Workbook Update
# -----------------------------
def update_tracker_workbook(wb, summary_df, raw_df):
    tracker_rows, known_tracker_parts = build_tracker_lookup(wb)

    matched = []
    not_found = []

    for _, row in summary_df.iterrows():
        item_no = recover_part_from_tracker(normalize_part_number(row["ItemNo"]), known_tracker_parts)
        qty = row["Quantity"]
        desc = row["Description"]
        pallet_list = row["PalletList"]
        document_list = row["DocumentList"]

        if item_no in tracker_rows:
            for entry in tracker_rows[item_no]:
                ws = entry["sheet"]
                excel_row = entry["row"]
                qty_col = entry["qty_col"]
                pallet_col = entry["pallet_col"]
                container_col = entry["container_col"]

                ws.cell(row=excel_row, column=qty_col).value = qty

                if pallet_col is not None:
                    existing_pallets = ws.cell(row=excel_row, column=pallet_col).value
                    ws.cell(row=excel_row, column=pallet_col).value = merge_comma_separated(existing_pallets, pallet_list)

                if container_col is not None:
                    existing_containers = ws.cell(row=excel_row, column=container_col).value
                    ws.cell(row=excel_row, column=container_col).value = merge_comma_separated(existing_containers, document_list)

                matched.append(
                    {
                        "ItemNo": item_no,
                        "Description": desc,
                        "Quantity": qty,
                        "PalletList": pallet_list,
                        "DocumentList": document_list,
                        "Sheet": ws.title,
                        "Row": excel_row,
                    }
                )
        else:
            not_found.append(
                {
                    "ItemNo": item_no,
                    "Description": desc,
                    "Quantity": qty,
                    "PalletList": pallet_list,
                    "DocumentList": document_list,
                }
            )

    unmatched_rows = []

    for _, row in raw_df.iterrows():
        item_no = recover_part_from_tracker(normalize_part_number(row["ItemNo"]), known_tracker_parts)

        if item_no not in tracker_rows:
            unmatched_rows.append(
                {
                    "Item #": row["ItemNo"],
                    "Normalized Item #": item_no,
                    "Description": row.get("Description", ""),
                    "Quantity": row.get("Quantity", ""),
                    "ColliNo": row.get("ColliNo", ""),
                    "DocumentNumber": row.get("DocumentNumber", ""),
                    "SourceFile": row.get("SourceFile", ""),
                    "PageNumber": row.get("PageNumber", ""),
                }
            )

    unmatched_df = pd.DataFrame(unmatched_rows)

    if not unmatched_df.empty:
        unmatched_df = (
            unmatched_df.groupby(
                ["Item #", "Normalized Item #", "Description", "ColliNo", "DocumentNumber", "SourceFile", "PageNumber"],
                as_index=False,
            )["Quantity"].sum()
        )

    matched_df = pd.DataFrame(matched)
    not_found_df = pd.DataFrame(not_found)

    return wb, matched_df, not_found_df, unmatched_df


# -----------------------------
# Export Helpers
# -----------------------------
def workbook_to_bytes(wb):
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def dataframe_to_excel_bytes(sheets_dict):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, df in sheets_dict.items():
            df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
    output.seek(0)
    return output.getvalue()


# -----------------------------
# UI
# -----------------------------
delivery_files = st.file_uploader(
    "Upload delivery PDFs/images",
    type=["pdf", "png", "jpg", "jpeg", "tif", "tiff"],
    accept_multiple_files=True,
    key=f"delivery_files_{st.session_state.run_id}",
)

tracker_file = st.file_uploader(
    "Upload tracker workbook",
    type=["xlsx", "xlsm"],
    key=f"tracker_file_{st.session_state.run_id}",
)

col_a, col_b = st.columns([1, 1])

with col_a:
    process_clicked = st.button("Process Files", type="primary")

with col_b:
    reset_clicked = st.button("Reset / Run New Files")

if reset_clicked:
    st.session_state.processed = False
    st.session_state.results = {}
    st.session_state.run_id += 1
    st.rerun()

if process_clicked:
    if not delivery_files:
        st.error("Please upload at least one delivery PDF/image.")
        st.stop()

    if not tracker_file:
        st.error("Please upload the tracker workbook.")
        st.stop()

    st.session_state.processed = False
    st.session_state.results = {}

    with st.spinner("Running OCR and updating tracker..."):
        tracker_bytes = tracker_file.getvalue()
        wb = load_workbook(BytesIO(tracker_bytes))

        _, known_tracker_parts = build_tracker_lookup(wb)

        raw_df, summary_df, preview_images, ocr_text_df, skipped_lines_df = process_delivery_files(
            delivery_files,
            known_tracker_parts
        )

        wb, matched_df, not_found_df, unmatched_df = update_tracker_workbook(wb, summary_df, raw_df)

        ocr_results_bytes = dataframe_to_excel_bytes(
            {
                "Parsed_OCR_Rows": raw_df if not raw_df.empty else pd.DataFrame(),
                "Summarized_Totals": summary_df if not summary_df.empty else pd.DataFrame(),
            }
        )

        parsed_rows_export_bytes = dataframe_to_excel_bytes(
            {"Parsed_OCR_Rows": raw_df if not raw_df.empty else pd.DataFrame()}
        )

        ocr_text_export_bytes = dataframe_to_excel_bytes(
            {"OCR_Raw_Text": ocr_text_df if not ocr_text_df.empty else pd.DataFrame()}
        )

        skipped_lines_export_bytes = dataframe_to_excel_bytes(
            {"Skipped_Lines": skipped_lines_df if not skipped_lines_df.empty else pd.DataFrame()}
        )

        updated_tracker_bytes = workbook_to_bytes(wb)

        unmatched_export_bytes = None
        if unmatched_df is not None and not unmatched_df.empty:
            unmatched_export_bytes = dataframe_to_excel_bytes({"Unmatched_OCR_Items": unmatched_df})

    st.session_state.results = {
        "raw_df": raw_df,
        "summary_df": summary_df,
        "matched_df": matched_df,
        "not_found_df": not_found_df,
        "unmatched_df": unmatched_df,
        "preview_images": preview_images,
        "updated_tracker_bytes": updated_tracker_bytes,
        "ocr_results_bytes": ocr_results_bytes,
        "parsed_rows_export_bytes": parsed_rows_export_bytes,
        "ocr_text_export_bytes": ocr_text_export_bytes,
        "ocr_text_df": ocr_text_df,
        "skipped_lines_df": skipped_lines_df,
        "skipped_lines_export_bytes": skipped_lines_export_bytes,
        "unmatched_export_bytes": unmatched_export_bytes,
    }

    st.session_state.processed = True
    st.rerun()

if st.session_state.processed:
    results = st.session_state.results

    raw_df = results["raw_df"]
    summary_df = results["summary_df"]
    matched_df = results["matched_df"]
    not_found_df = results["not_found_df"]
    unmatched_df = results["unmatched_df"]
    preview_images = results["preview_images"]
    updated_tracker_bytes = results["updated_tracker_bytes"]
    ocr_results_bytes = results["ocr_results_bytes"]
    parsed_rows_export_bytes = results["parsed_rows_export_bytes"]
    ocr_text_export_bytes = results["ocr_text_export_bytes"]
    ocr_text_df = results["ocr_text_df"]
    skipped_lines_df = results["skipped_lines_df"]
    skipped_lines_export_bytes = results["skipped_lines_export_bytes"]
    unmatched_export_bytes = results["unmatched_export_bytes"]

    st.success("Processing complete.")

    if preview_images:
        with st.expander("Preview OCR Images", expanded=False):
            for name, img in preview_images:
                st.markdown(f"**{name}**")
                st.image(img, use_container_width=True)

    col1, col2 = st.columns(2)

    with col1:
        st.subheader("Summarized Totals")
        st.dataframe(summary_df, use_container_width=True)

    with col2:
        st.subheader("Matched Rows")
        st.dataframe(matched_df, use_container_width=True)

    with st.expander("Parsed OCR Rows Preview", expanded=False):
        st.dataframe(raw_df, use_container_width=True)

    with st.expander("OCR Raw Text Preview", expanded=False):
        st.dataframe(ocr_text_df, use_container_width=True)

    with st.expander("Skipped Lines Preview", expanded=False):
        if not skipped_lines_df.empty:
            st.dataframe(skipped_lines_df, use_container_width=True)
        else:
            st.info("No skipped lines were recorded.")

    st.subheader("Unmatched Items")
    if unmatched_df is not None and not unmatched_df.empty:
        st.dataframe(unmatched_df, use_container_width=True)
    else:
        st.info("No unmatched items.")

    if not not_found_df.empty:
        with st.expander("Summary items not found in tracker", expanded=False):
            st.dataframe(not_found_df, use_container_width=True)

    st.download_button(
        label="Download Updated Tracker",
        data=updated_tracker_bytes,
        file_name="Updated_Tracking_File.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.download_button(
        label="Download OCR Results",
        data=ocr_results_bytes,
        file_name="OCR_Results.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.download_button(
        label="Download Parsed OCR Rows",
        data=parsed_rows_export_bytes,
        file_name="Parsed_OCR_Rows.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.download_button(
        label="Download OCR Raw Text",
        data=ocr_text_export_bytes,
        file_name="OCR_Raw_Text.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.download_button(
        label="Download Skipped Lines",
        data=skipped_lines_export_bytes,
        file_name="Skipped_Lines.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    if unmatched_export_bytes is not None:
        st.download_button(
            label="Download Unmatched Items",
            data=unmatched_export_bytes,
            file_name="Unmatched_OCR_Items.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
else:
    st.info("Upload your files and click Process Files. The app will only rerun processing when you press that button or reset the session.")
